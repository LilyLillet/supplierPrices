import openpyxl


def write_to_excel(lst, file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    rows = sheet.max_row
    for row in range(1, rows + 1):
        good = str(sheet.cell(row=row, column=1).value)
        n = good.lower()
        for i in lst:
            col_value = int(i[0])
            s = i[1:]
            g = i.split()
            ln = len(g)
            price = g[ln - 1]
            c = s.replace(" " + price, "")
            o = c.lower()
            if o in n:
                cell = sheet.cell(row=row, column=col_value)
                cell.value = price
    workbook.save('gorb.xlsx')
